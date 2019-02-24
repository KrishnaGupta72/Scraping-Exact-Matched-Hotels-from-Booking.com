import openpyxl
import time
from selenium import webdriver
import csv
# from Check_Rubbish import strcheckrubbish

path = "E:\\PyCharm Projects\\Input_File.xlsx"

wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row=1, column=1)
input_hotel_list = []
input_city_list = []
input_country_list = []
new_map_zone_urls_list = ['Url_Col', 'index_val']
max_row = sheet_obj.max_row
for i in range(2, (max_row + 1)):
    cell_obj = sheet_obj.cell(row=i, column=5)#5
    input_hotel_list.append(cell_obj.value)
    cell_obj = sheet_obj.cell(row=i, column=7)#8
    input_city_list.append(cell_obj.value)
    cell_obj = sheet_obj.cell(row=i, column=6)#7
    input_country_list.append(cell_obj.value)

driver = webdriver.Chrome("E:\\PyCharm Projects\\chromedriver.exe")
with open("Output_Hotel_Matching.csv", "w", encoding='utf-8', newline='') as file:

    field_names = ['Input_HotelName', 'Input_CityName', 'Input_CountryName', 'Output_HotelName',
                   'Output_HotelName_Address']
    writer = csv.DictWriter(file, fieldnames=field_names)
    writer.writeheader()
    for count_4_header, (hotel, city, country) in enumerate(zip(input_hotel_list, input_city_list, input_country_list)):
        hotel_name = hotel
        hotel_city_country = hotel + ' ' + city + ' ' + country
        hotel_city = hotel + ' ' + city
        driver.get('https://www.booking.com/')
        time.sleep(1)

        # Condition for Hotel_Name search
        search_box = driver.find_element_by_xpath(
            '//input[@class="c-autocomplete__input sb-searchbox__input sb-destination__input"]')
        search_box.clear()

        search_box.send_keys(hotel_name)#hotel_city_country #hotel_city #hotel_name
        time.sleep(1)
        # ##################write a response file after passing a Hotel's name ####################
        # Location_mapping_Resp = driver.page_source
        # Loc_map = open("Location_mapping" + hotel + ".html", "w", encoding="utf-8")
        # Loc_map.write(Location_mapping_Resp)
        ###########################################################
        # Storing suggested Hotel's element into a list.

        Hotel_list = driver.find_elements_by_xpath(
            '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span/b[@class="search_hl_name"][1]')
        Hotel_list_with_address = driver.find_elements_by_xpath(
            '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span')

        #Condition for Hotel,City and Country search
        if len(Hotel_list)==0:
            search_box = driver.find_element_by_xpath(
                '//input[@class="c-autocomplete__input sb-searchbox__input sb-destination__input"]')
            search_box.clear()

            search_box.send_keys(hotel_city_country)  # hotel_city_country #hotel_city #hotel_name
            time.sleep(1)

            Hotel_list = driver.find_elements_by_xpath(
                '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span/b[@class="search_hl_name"][1]')
            Hotel_list_with_address = driver.find_elements_by_xpath(
                '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span')

        # Condition for Hotel and City search
        if len(Hotel_list) == 0:
            search_box = driver.find_element_by_xpath(
                '//input[@class="c-autocomplete__input sb-searchbox__input sb-destination__input"]')
            search_box.clear()

            search_box.send_keys(hotel_city)  # hotel_city_country #hotel_city #hotel
            time.sleep(1)

            Hotel_list = driver.find_elements_by_xpath(
                '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span/b[@class="search_hl_name"][1]')
            Hotel_list_with_address = driver.find_elements_by_xpath(
                '//li[@class="c-autocomplete__item sb-autocomplete__item sb-autocomplete__item-with_photo sb-autocomplete__item--hotel sb-autocomplete__item__item--elipsis "]/span')

        for hotel_name, Hotel_add in zip(Hotel_list, Hotel_list_with_address):
            #Replacing special characters for City and Country Comparision.
            city = city.replace("'", "'")
            site_Hotel_add=Hotel_add.text.replace("'", "'")

            #Condition to handle country "Italy" as a "Italien".
            if country=='Italy':
                temp_country = "Italien"
                if (city.upper() in site_Hotel_add.upper()) and ((country.upper() in site_Hotel_add.upper()) or (temp_country.upper() in site_Hotel_add.upper())):
                    try:
                        writer.writerow(
                            {
                                'Input_HotelName': hotel,
                                'Input_CityName': city,
                                'Input_CountryName': country,
                                'Output_HotelName': hotel_name.text,
                                'Output_HotelName_Address': site_Hotel_add

                            }
                        )
                    except:
                        continue



            elif (city.upper() in site_Hotel_add.upper()) and (country.upper() in site_Hotel_add.upper()):

                try:
                    writer.writerow(
                        {
                            'Input_HotelName': hotel,
                            'Input_CityName': city,
                            'Input_CountryName': country,
                            'Output_HotelName': hotel_name.text,
                            'Output_HotelName_Address': site_Hotel_add

                        }
                    )
                except:
                    continue
